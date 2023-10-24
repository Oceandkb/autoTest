# @Time   ：2023/9/6 9:32
# @Author : Chao
# @File   : test_question_add.py


import allure
from aikn_api.question_add import question_add
import requests
from openpyxl import load_workbook

@allure.epic("知识库")
@allure.feature("社区")
@allure.story("新增问题成功")
def test_question_add(login_fixture, base_url, community_plate_search_fixture, datetime_fixture):

    # 打开 Excel 文件，假设文件名为 data.xlsx
    workbook = load_workbook("testcases/community/question.xlsx")
    sheet = workbook.worksheets[0]  # 选择第一个工作表

    # 循环遍历第一列数据
    q = []
    for row in sheet.iter_rows(values_only=True):
        q.append(row[0])
        #que = ''.join(q)
    # print(que, type(que))

    for que in q:
        r = question_add(login_fixture, base_url, question_title = que + datetime_fixture, question_content = '<p>' + que + datetime_fixture + '</p>',
                         plate_id = community_plate_search_fixture)
        # print(community_plate_search_fixture, type(community_plate_search_fixture))
        print(r)
        assert r.json()["code"] == 1
        assert r.status_code == 200
        assert r.json()['message'] == '操作成功'