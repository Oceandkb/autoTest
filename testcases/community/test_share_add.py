# @Time   ：2023/9/12 17:24
# @Author : Chao
# @File   : test_share_add.py


import allure
from aikn_api.share_add import share_add
import requests
from openpyxl import load_workbook

@allure.epic("知识库")
@allure.feature("社区")
@allure.story("新增分享成功")
def test_share_add(login_fixture, base_url, community_plate_search_fixture):

    # 打开 Excel 文件，假设文件名为 data.xlsx
    workbook = load_workbook("question.xlsx")
    sheet = workbook.worksheets[1]  # 选择第一个工作表

    # 创建列表，存储键值对
    s_list = []
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        key = row[0]
        value = row[1]
        s_list.append([key, value])
        #que = ''.join(q)
#print(s_list)

    for s in s_list:
        print(s, type(s))
        key = s[0]
        value = s[1]
        r = share_add(login_fixture, base_url, share_title = key, share_content='<p>'+ value +'</p>', class_id='-1',
                  plate_id=community_plate_search_fixture)
    print(r)
    assert r.json()["code"] == 1
    assert r.status_code == 200
    assert r.json()['message'] == '操作成功'