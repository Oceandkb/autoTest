# @Time   ：2023/8/21 14:53
# @Author : Chao
# @File   : demo1.py

from openpyxl import load_workbook

# 打开 Excel 文件，假设文件名为 data.xlsx
workbook = load_workbook("demo1.xlsx")
sheet = workbook.worksheets[0]  # 选择第一个工作表


# 循环遍历第一列的数据
column_data = []
column_data1 = []
for row in sheet.iter_rows(values_only=True):
    column_data.append(row[0])
    print(type(column_data))
